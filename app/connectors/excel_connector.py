from openpyxl import load_workbook


def load_excel(file_path):
    """
    Загружает Excel-файл и возвращает все строки.
    """

    workbook = load_workbook(file_path, data_only=True)

    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    return rows


def get_headers(file_path):
    """
    Возвращает список заголовков первой строки Excel.
    """

    workbook = load_workbook(file_path, read_only=True)

    sheet = workbook.active

    headers = next(sheet.iter_rows(values_only=True))

    return list(headers)


def detect_excel_type(file_path):
    """
    Определяет тип Excel-файла по заголовкам.
    """

    headers = get_headers(file_path)

    if "Всего регистраций" in headers:
        return "registrations"

    if "Сумма в валюте отчета" in headers:
        return "deposits"
    conversion_headers = {
        "Агент",
        "Агент (ID)",
        "Субагент",
        "Субагент (ID)",
        "Страна",
        "Количество 'OK'",
        "Доля 'OK', %",
        "Общее количество заявок",
    }

    if conversion_headers.issubset(set(headers)):
        return "conversions"
    
    return None


def find_country(rows):
    """
    Ищет колонку Страна/Country и берет значение из второй строки.
    """

    headers = rows[0]

    country_column = None

    for index, header in enumerate(headers):
        if header in ["Страна", "Country"]:
            country_column = index
            break

    if country_column is None:
        raise Exception("Не найдена колонка Страна/Country")

    return rows[1][country_column]